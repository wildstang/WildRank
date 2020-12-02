# csv2xls.py
# Liam Fruzyna 2020-05-21
# Convert a LiamRank CSV file export to an Excel Spreadsheet with Sheets

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import sys, string

# read in the given csv file
if len(sys.argv) < 2:
    sys.exit('No CSV file provided, exiting')
df = pd.read_csv(sys.argv[1])

if len(sys.argv) < 3:
    # create workbook and sheets
    wb = Workbook()
    wb.create_sheet('pit')
    wb.create_sheet('match')
    wb.create_sheet('note')
else:
    # load in existing workbook
    wb = load_workbook(sys.argv[2])

# add each row of csv to workbook
for row in dataframe_to_rows(df, index=False, header=True):
    kind = row[2]
    if kind != 'kind':
        # adetermine sheet to use
        if kind == 'pit':
            ws = wb['pit']
        elif kind == 'match':
            ws = wb['match']
        elif kind == 'note':
            ws = wb['note']
        else:
            continue

        # check if row already exists
        unique = True
        for r in ws.iter_rows(min_row=2):
            if r[0].value == row[0]:
                unique = False

                # check if existing version is identica;
                identical = True
                for i in range(len(r)):
                    if r[i].value != row[i]:
                        identical = False
                        break

                if identical:
                    # skip if identical
                    print(row[0], 'already exists, skipping')
                else:
                    # prompt to replace if not
                    print('found 2 different versions of row', row[0])
                    print('old:', [a.value for a in r])
                    print('new:', row)
                    if input('Replace existing row? [y/N] ').upper() == 'Y':
                        for i in range(len(r)):
                            r[i].value = row[i]
        
        # add row if unique
        if unique:
            ws.append(row)
    elif len(sys.argv) < 3:
        # add header to each sheet
        wb['pit'].append(row)
        wb['match'].append(row)
        wb['note'].append(row)

# remove empty columns from sheets
for sheet in ['pit', 'match', 'note']:
    i = 1
    while i <= len(wb[sheet][1]):
        # get list of column names
        cols = list(map(lambda cell: cell.value, list(wb[sheet][1])))
        # get all non-nan cells
        cells = [cell.value for cell in wb[sheet][get_column_letter(i)] if str(cell.value) != 'nan']
        # remove header
        name = cells.pop(0)
        if len(cells) == 0:
            print('Deleting empty column from', sheet, '-', name)
            wb[sheet].delete_cols(cols.index(name)+1)
        else:
            i += 1

# write workbook to file
wb.save('export.xlsx')