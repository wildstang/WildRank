# csv2xls.py
# Liam Fruzyna 2020-05-21
# Convert a LiamRank CSV file export to an Excel Spreadsheet with Sheets

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import sys

# read in the given csv file
if len(sys.argv) < 2:
    sys.exit('No CSV file provided, exiting')
df = pd.read_csv(sys.argv[1])

if len(sys.argv) < 3:
    # create workbook and sheets
    wb = Workbook()
    wb.create_sheet('pit')
    wb.create_sheet('match')
    wb.create_sheet('notes')
else:
    # load in existing workbook
    wb = load_workbook(sys.argv[2])

# add each row of csv to workbook
for row in dataframe_to_rows(df, index=False, header=True):
    kind = row[2]
    if kind != 'kind':
        # adetermine sheet to use
        if kind == 'pit':
            ws = wb['pit']
        elif kind == 'match':
            ws = wb['match']
        elif kind == 'notes':
            ws = wb['notes']
        else:
            continue

        # check if row already exists
        unique = True
        for r in ws.iter_rows(min_row=2):
            if r[0].value == row[0]:
                unique = False

                # check if existing version is identica;
                identical = True
                for i in range(len(r)):
                    if r[i].value != row[i]:
                        identical = False
                        break

                if identical:
                    # skip if identical
                    print(row[0], 'already exists, skipping')
                else:
                    # prompt to replace if not
                    print('found 2 different versions of row', row[0])
                    print('old:', [a.value for a in r])
                    print('new:', row)
                    if input('Replace existing row? [y/N] ').upper() == 'Y':
                        for i in range(len(r)):
                            r[i].value = row[i]
        
        # add row if unique
        if unique:
            ws.append(row)
    elif len(sys.argv) < 3:
        # add header to each sheet
        wb['pit'].append(row)
        wb['match'].append(row)
        wb['notes'].append(row)

# write workbook to file
wb.save('export.xlsx')